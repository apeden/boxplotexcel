from openpyxl import load_workbook
import numpy as np
import matplotlib.pyplot as plt
import pylab
from scipy import stats


workbooks = ["selectAll", "selectAllNegativeInteresting"]

workbook = "C:\\Users\\user\\OneDrive - University of Edinburgh\\excelSheet\\boxplotexcel\\age_at_death.xlsx"

class Sheet(object):
    def __init__(self, excel_workbook_name, sheet_name):
                #clear previous workbook
        self.wb = None
        self.excel_workbook_name = excel_workbook_name
        self.sheet_name = sheet_name
        ##set workbook as file object
        try:
            self.wb = load_workbook(self.excel_workbook_name, data_only = True)
        except IOError:
            self.wb = None
            print("Could not open file"+ self.excel_workbook_name)
        ##get sheet as file object
        self.sheet = self.wb[self.sheet_name]
        
class Column(Sheet):
    """for generating box plots from data in columns"""
    def __init__(self, excel_workbook_name, sheet_name):
        Sheet.__init__(self, excel_workbook_name, sheet_name)
        self.column_label = ""
        self.column_list = []
        self.mult_column_lists = []
    def set_column_label(self, row, column):
        try:
            val = self.sheet.cell(row, column).value
        except:   
            raise ColumnLabelError("A problem occurred setting the column label.")
        self.column_label = val     
    def get_column_label(self):
        return self.column_label
    def set_column_list(self, start_row, column):
        self.column_list = []
        for row in range(start_row, 1000):        
            try:
                val = self.sheet.cell(row, column).value
            except AttributeError:   
                print("A problem occurred reading data from a row")            
            if not val == None:
                self.column_list = self.column_list + [float(val)]    
    def get_column_list(self):
        return self.column_list
    def set_mult_column_lists(self):
        self.mult_column_lists = []
        i = 1
        while True:
            self.set_column_list(2, i)
            if self.column_list == []:
                break;
            self.mult_column_lists.append(self.column_list)
            i += 1
    def get_mult_column_lists(self):
        return self.mult_column_lists
    
        #remember plt = matplotlib.pyplot

def display_box_plot(variable, column_list):
    fig, box_plot = plt.subplots() 
    box_plot.set_title(variable)
    box_plot.boxplot(column_list)
    pylab.show()

def display_mult_box_plots(variable, mult_column_lists):
    fig, box_plot = plt.subplots()
    box_plot.set_title(variable)
    box_plot.set_ylabel(variable)
    box_plot.boxplot(mult_column_lists)
    box_plot.set_xticklabels(workbooks, rotation=45, fontsize=8)
    pylab.show()


def inter_sheet_column_box_plot(column):
    columns =[]
    label = ""
    for workbook in workbooks:
        c = Column(workbook + ".xlsx","selectAll") 
        if label is "":
            c.set_column_label(1,column)
            label = c.get_column_label()
        c.set_column_list(2,column)
        columns.append(c.get_column_list())
    display_mult_box_plots(label, columns)

#for column in range(1,7):
   # try:
   #     inter_sheet_column_box_plot(column)
   # except:
   #     print("Could not plot box plots for column "+ str(column))

def inter_sheet_column_ttest(column):
    columns =[]
    label = ""
    for workbook in workbooks:
        c = Column(workbook + ".xlsx","selectAll") 
        if label is "":
            c.set_column_label(1,column)
            label = c.get_column_label()
        c.set_column_list(2,column)
        columns.append(c.get_column_list())
    return stats.ttest_ind(columns[0],columns[1], equal_var = False), label

for column in range(1,7):
    try:
        print(inter_sheet_column_ttest(column))
    except:
        print("Could not perform tttest for "+ str(column))

##print(c.get_mult_column_lists())
##c.set_column_list(2,1)
##print(c.get_column_list())
##c.display_mult_box_plots("AgeAtDeath")
##c.display_box_plot("AgeAtDeath")

    


### method for creating work book object and worksheetobject
### method for extracting data from a column and putting it in a list
###set_column_list
###get_column_list
###method for setting up plot
